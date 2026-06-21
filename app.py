# ============================================================
# BITÁCORA DIGITAL DE MANTENIMIENTO AEP
# Aplicación Streamlit para registro de actividades de mantenimiento
# ============================================================

import streamlit as st
import gspread
import pandas as pd
import unicodedata
from google.oauth2.service_account import Credentials
from datetime import datetime, date
from io import BytesIO
import base64
import altair as alt
import os
import pytz
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, PageBreak, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO



st.markdown(
    """
    <link rel="manifest" href="/manifest.json">
    <meta name="theme-color" content="#0E1117">
    """,
    unsafe_allow_html=True
)

st.set_page_config(page_title="Bitácora Digital de Mantenimiento", page_icon="🛠", layout="wide")

st.markdown("""
<style>
[data-testid="stSidebar"] img {border-radius:50%;display:block;margin-left:auto;margin-right:auto;}
.login-box{
    background:linear-gradient(180deg,rgba(0,0,0,0.78),rgba(25,25,25,0.88));

    padding:50px 60px;

    border-radius:22px;

    max-width:800px;   /* 🔥 MÁS ANCHO */
    width:95%;

    margin:60px auto;

    color:white;
    box-shadow:0px 25px 60px rgba(0,0,0,0.6);

    height:auto;
    overflow:visible;
}
.login-title{
    font-size:44px;
    font-weight:900;        /* 🔥 MÁS GRUESO */
    text-align:center;
    letter-spacing:1.5px;   /* 🔥 MÁS INDUSTRIAL */
    line-height:1.2;
    
    text-transform:uppercase;
    
    text-shadow: 0px 2px 10px rgba(0,0,0,0.6);  /* 🔥 PROFUNDIDAD */
}
.login-subtitle{font-size:22px;text-align:center;margin-top:10px;font-weight:500;}
.login-line{margin:18px auto;width:80px;height:3px;background:#2ecc71;border-radius:10px;}
.login-features{text-align:center;font-size:14px;opacity:0.85;margin-bottom:25px;}
.login-footer{text-align:center;font-size:11px;opacity:0.5;margin-top:30px;}
            @keyframes girar {
    from { transform: rotate(0deg); }
    to { transform: rotate(360deg); }
}

.gear {
    font-size:50px;
    text-align:center;
    margin-bottom:15px;
    animation: girar 6s linear infinite;
}
</style>

""", unsafe_allow_html=True)

# ================= GOOGLE =================
scope=["https://spreadsheets.google.com/feeds","https://www.googleapis.com/auth/drive"]
credentials=Credentials.from_service_account_info(st.secrets["google_credentials"],scopes=scope)
gc=gspread.authorize(credentials)
sheet=gc.open("Bitacora_Mantenimiento")
ws_usuarios=sheet.worksheet("Usuarios")
ws_ots=sheet.worksheet("OTs")
ws_bitacora=sheet.worksheet("Bitacora")
ws_adicional = sheet.worksheet("Actividades_Adicionales")

if "login" not in st.session_state: st.session_state.login=False
if "area" not in st.session_state: st.session_state.area=None

# ================= ROLES =================
ROLES_TECNICOS = [
    "MECÁNICO",
    "ELECTRICISTA",
    "INSTRUMENTISTA",
    "RECORREDOR DE DUCTOS"
]

def portada_login(image_file):
    with open(image_file,"rb") as f:
        encoded=base64.b64encode(f.read()).decode()
    st.markdown(
        f"<style>.stApp{{background-image:url('data:image/jpg;base64,{encoded}');background-size:cover;background-position:center;}}</style>",
        unsafe_allow_html=True
    )

# ================= LOGIN =================
if not st.session_state.login:
    portada_login("fondo_planta.jpg")
    st.markdown("""
    <div class="login-box">
        <div class="gear">⚙️</div>
        <div class="login-title">SISTEMA DE GESTIÓN DE MANTENIMIENTO</div>
        <div class="login-subtitle">MECÁNICO · EI&C · GIA · EHS</div>
        <div class="login-line"></div>
        <div class="login-features">
            Planeamiento · Ejecución · Confiabilidad · KPIs<br>
            ✔ Control de OTs ✔ Actividades no planificadas ✔ Análisis de fallas
        </div>
    """,unsafe_allow_html=True)
    usuario=st.text_input("USUARIO")
    password=st.text_input("CONTRASEÑA",type="password")
    if st.button("INGRESAR"):
        df_users=pd.DataFrame(ws_usuarios.get_all_records())
        valid=df_users[(df_users["Usuario"]==usuario)&(df_users["Password"].astype(str)==password)]
        if not valid.empty:
            st.session_state.login=True
            st.session_state.usuario=usuario
            st.session_state.nombre=valid.iloc[0]["Nombre"]
            st.session_state.rol=valid.iloc[0]["Rol"]
            st.session_state.area=valid.iloc[0]["area"]
            st.session_state.cargo = valid.iloc[0].get("cargo", st.session_state.rol)
            st.rerun()
        else:
            st.error("Credenciales incorrectas")
    st.markdown('<div class="login-footer">Sistema interno de gestión de mantenimiento · v1.0</div></div>',unsafe_allow_html=True)
    st.stop()

# ================= SIDEBAR =================
ruta_foto=f"fotos/{st.session_state.usuario}.jpg"
ruta_default="fotos/default.jpg"
with st.sidebar:
    if os.path.exists(ruta_foto): st.image(ruta_foto)
    elif os.path.exists(ruta_default): st.image(ruta_default)
    else: st.markdown("👤 Sin foto")
st.sidebar.success(st.session_state.nombre)
st.sidebar.info(st.session_state.cargo)
st.sidebar.info(f"Área: {st.session_state.area}")
if st.sidebar.button("Cerrar sesión"):
    st.session_state.clear()
    st.rerun()

# ================= BITÁCORA (DURACIÓN AUTOMÁTICA) =================
if st.session_state.rol in ROLES_TECNICOS:
    st.title(f"📝BITÁCORA DIARIA – {st.session_state.area}")

    tab_registro, tab_adicional, tab_mis_registros, tab_avance = st.tabs([
        "📝 Registrar OT",
        "➕ Actividad adicional",
        "✏️ Editar registros",
        "📈 Mi avance"
    ])

    # ================= REGISTRAR OT =================
    with tab_registro:
        df_plan = pd.DataFrame(ws_ots.get_all_records())
        df_plan.columns = df_plan.columns.str.strip().str.lower()
        if "fecha ejecucion" not in df_plan.columns:
            st.error("❌ La hoja OTs no tiene la columna 'fecha ejecucion'")
            st.stop()
        df_plan["fecha ejecucion"] = pd.to_datetime(
            df_plan["fecha ejecucion"], 
            dayfirst = True,
            errors="coerce"
        ).dt.date
        df_plan["area"] = df_plan["area"].astype(str).str.strip()
        df_plan = df_plan[df_plan["area"] == st.session_state.area]

        fecha_sel = st.date_input("Fecha", value=date.today())

        df_hoy = df_plan[df_plan["fecha ejecucion"] == fecha_sel]

        df_bit = pd.DataFrame(ws_bitacora.get_all_records())
        df_plan.columns = df_plan.columns.str.strip().str.lower()
        if "fecha ejecucion" not in df_plan.columns:
            st.error("❌ La hoja OTs no tiene la columna 'fecha'")
            st.stop()
        df_bit["fecha"] = pd.to_datetime(df_bit["fecha"], errors="coerce").dt.date

        # Asegurar columnas críticas
        df_bit["pt"] = df_bit["pt"].astype(str).str.strip().str.upper()
        df_bit["ot"] = df_bit["ot"].astype(str).str.strip().str.upper()
        
        # ===== ASEGURAR COLUMNA actividad_plan =====
        if "actividad_plan" not in df_bit.columns:
            df_bit["actividad_plan"] = ""
        
        df_bit["actividad_plan"] = df_bit["actividad_plan"].astype(str)

        df_bit["pt"] = df_bit["pt"].astype(str).str.strip()

        #SI PT ≠ S/PT  → consumir por PT
        #SI PT = S/PT Y OT ≠ S/OT → consumir por OT
        #SI PT = S/PT Y OT = S/OT → consumir por ACTIVIDAD
        # ===== NORMALIZAR COLUMNA DE ACTIVIDAD PARA CONSUMO =====
        
        df_bit["actividad_consumo"] = df_bit["detalle"].astype(str)
        df_hoy["actividad_consumo"] = df_hoy["actividad"].astype(str)

        # ================= CLAVE DE CONSUMO DIARIA =================
        def normalizar_texto(txt):
            txt = str(txt).upper().strip()
            txt = unicodedata.normalize("NFKD", txt)
            txt = "".join(c for c in txt if not unicodedata.combining(c))
            txt = (
                txt.replace("  ", " ")
                   .replace(".", "")
                   .replace(",", "")
            )
            return txt
        
        # --- NORMALIZACIÓN ---

        for col in ["pt", "ot"]:
            df_bit[col] = df_bit[col].astype(str).str.strip().str.upper()
            df_hoy[col] = df_hoy[col].astype(str).str.strip().str.upper()
        
        # IMPORTANTE:
        # Bitácora usa "actividad_plan"
        # Planeamiento usa "actividad"

        df_bit["actividad_plan_norm"] = df_bit["actividad_plan"].apply(normalizar_texto)
        df_hoy["actividad_plan_norm"] = df_hoy["actividad"].apply(normalizar_texto)

        # --- CLAVE DE CONSUMO ---

        def clave_consumo(row):
            if row["pt"] not in ["S/PT", "", "NONE"]:
                return f"PT_{row['pt']}"
            elif row["ot"] not in ["S/OT", "", "NONE"]:
                return f"OT_{row['ot']}"
            else:
                # Caso S/PT + S/OT → por actividad
                return f"ACT_{row['actividad_plan_norm']}"
       
        df_bit["actividad_plan_norm"] = df_bit["actividad_plan"].apply(normalizar_texto)
        df_hoy["actividad_plan_norm"] = df_hoy["actividad"].apply(normalizar_texto)

        df_bit["clave_consumo"] = df_bit.apply(clave_consumo, axis=1)
        df_hoy["clave_consumo"] = df_hoy.apply(clave_consumo, axis=1)
    
        # ================= CONSUMOS YA REGISTRADOS HOY =================
        df_consumidos = df_bit[
            (df_bit["area"] == st.session_state.area) &
            (df_bit["fecha"] == fecha_sel)
        ].copy()

        # Asegurar normalización
        df_consumidos["pt"] = df_consumidos["pt"].astype(str).str.upper().str.strip()
        df_consumidos["ot"] = df_consumidos["ot"].astype(str).str.upper().str.strip()
        df_consumidos["actividad_plan_norm"] = df_consumidos["actividad_plan"].apply(normalizar_texto)

        # Construir clave_consumo REAL (bitácora)
        def clave_consumo_bit(row):
            if row["pt"] not in ["S/PT", "", "NONE"]:
                return f"PT_{row['pt']}"
            elif row["ot"] not in ["S/OT", "", "NONE"]:
                return f"OT_{row['ot']}"
            else:
                return f"ACT_{row['actividad_plan_norm']}"

        df_consumidos["clave_consumo"] = df_consumidos.apply(clave_consumo_bit, axis=1)

        # --- Planeamiento ---
        df_hoy["pt"] = df_hoy["pt"].astype(str).str.upper().str.strip()
        df_hoy["ot"] = df_hoy["ot"].astype(str).str.upper().str.strip()
        df_hoy["actividad_plan_norm"] = df_hoy["actividad"].apply(normalizar_texto)

        def clave_consumo_plan(row):
            if row["pt"] not in ["S/PT", "", "NONE"]:
                return f"PT_{row['pt']}"
            elif row["ot"] not in ["S/OT", "", "NONE"]:
                return f"OT_{row['ot']}"
            else:
                return f"ACT_{row['actividad_plan_norm']}"
        
        df_hoy["clave_consumo"] = df_hoy.apply(clave_consumo_plan, axis=1)

        # ================= FILTRO DEFINITIVO DIARIO =================
        # Normalizar actividad en ambos lados
        df_hoy["actividad_canon"] = df_hoy["actividad"].apply(normalizar_texto)
        df_bit["actividad_canon"] = df_bit["actividad_plan"].apply(normalizar_texto)

        df_consumidos = df_bit[
            (df_bit["area"] == st.session_state.area) &
            (df_bit["fecha"] == fecha_sel)
        ].copy()

        # ---------- CASO 1: PT ----------
        pts_consumidos = df_consumidos[
            df_consumidos["pt"] != "S/PT"
        ]["pt"].unique()

        df_hoy = df_hoy[~df_hoy["pt"].isin(pts_consumidos)]

        # ---------- CASO 2: OT ----------
        ots_consumidos = df_consumidos[
            (df_consumidos["pt"] == "S/PT") &
            (df_consumidos["ot"] != "S/OT")
        ]["ot"].unique()

        df_hoy = df_hoy[~df_hoy["ot"].isin(ots_consumidos)]

        # ---------- CASO 3: S/PT + S/OT → ACTIVIDAD ----------
        actividades_consumidas = df_consumidos[
            (df_consumidos["pt"] == "S/PT") &
            (df_consumidos["ot"] == "S/OT")
        ]["actividad_canon"].unique()

        df_hoy = df_hoy[
        ~df_hoy["actividad_canon"].isin(actividades_consumidas)
        ]

        if df_hoy.empty:
            st.success("✅ Ya registraste todas tus PTs del día")
            
        else:
            ot_sel = st.selectbox("PT", df_hoy["pt"].tolist())
            fila = df_hoy[df_hoy["pt"] == ot_sel].iloc[0]
            
            # ===== HISTÓRICO (TODO JUNTO) =====
            df_hist = pd.DataFrame(ws_bitacora.get_all_records())
            df_hist["avance_dia"] = pd.to_numeric(df_hist["avance_dia"], errors="coerce")
            df_hist["fecha"] = pd.to_datetime(df_hist["fecha"], errors="coerce").dt.date
        # ================= AVANCE PREVIO INTELIGENTE =================
       
            if "actividad_plan" not in df_hist.columns:
                df_hist["actividad_plan"] = ""

            df_hist["actividad_plan"] = df_hist["actividad_plan"].astype(str)

            df_hist_filtrado = df_hist[
                (df_hist["area"] == st.session_state.area) &
                (df_hist["fecha"] < fecha_sel)
            ].copy()

        # Normalizar columnas del histórico (bitácora)
            for col in ["pt", "ot", "detalle"]:
                df_hist_filtrado[col] = (
                    df_hist_filtrado[col].astype(str).str.strip().str.upper()
                )
        # Normalizar fila actual (OTs)
            for col in ["pt", "ot", "actividad"]:
                fila[col] = str(fila[col]).strip().upper()

        # ---------- CRITERIOS ----------
            if fila["pt"] not in ["S/PT", "", "NONE"]:
            # CASO 1: PT válida → consumir por PT
                avance_prev =df_hist_filtrado[
                    df_hist_filtrado["pt"] == fila["pt"]
                ]["avance_dia"].max()

            elif fila["ot"] not in ["S/OT", "", "NONE"]:
            # CASO 2: S/PT + OT → consumir por OT
                avance_prev = df_hist_filtrado[
                    df_hist_filtrado["ot"] == fila["ot"]
                ]["avance_dia"].max()
            
            else:
            # CASO 3: S/PT + S/OT → consumir por ACTIVIDAD
                actividad_norm = normalizar_texto(fila["actividad"])
                avance_prev = df_hist_filtrado[
                    df_hist_filtrado["actividad_plan"].apply(normalizar_texto) == actividad_norm
                ]["avance_dia"].max()    
    

            if pd.isna(avance_prev):
                avance_prev = 0

            df_users = pd.DataFrame(ws_usuarios.get_all_records())
            recursos = df_users[df_users["area"] == st.session_state.area]["Nombre"].tolist()
            recursos.insert(0, "N/A")

            with st.form("bitacora", clear_on_submit=True):
                # ===== FILA 1: OT | PT =====
                col1, col2 = st.columns(2)
                with col1:
                    st.text_input("OT", fila["ot"], disabled=True)
                with col2:
                    st.text_input("PT", fila["pt"], disabled=True)

                # ===== FILA 2: ACTIVIDAD =====
                st.text_area(
                    "Actividad",
                    fila["actividad"],
                    disabled=True,
                    height=90
                )
                # ===== FILA 3: DETALLE EJECUTADO =====
                detalle = st.text_area(
                    "Detalle ejecutado",
                    height=120
                )
                from datetime import time
                horas_turno = (
                    [time(h, 0) for h in range(7, 12)] +
                    [time(12, 0)] +
                    [time(13, 30)] +
                    [time(h, 0) for h in range(14, 20)]
                )
                # ===== FILA 4: HORA INICIO | HORA CIERRE =====
                col3, col4 = st.columns(2)
                with col3:
                    hora_inicio = st.selectbox("Hora inicio", horas_turno)
                with col4:
                    hora_cierre = st.selectbox("Hora cierre", horas_turno)

                # ===== FILA 5: RECURSO | AVANCE =====
                col5, col6 = st.columns(2)
                with col5:
                    recurso = st.selectbox("Recurso personal (apoyo)", recursos)
                with col6:
                    if avance_prev >= 100:
                        st.info("✅ Esta OT ya alcanzó el 100% de avance")
                        avance = 100
                    else:
                        avance = st.slider(
                            "Avance acumulado de la OT (%)",
                            min_value=int(avance_prev),
                            max_value=100,
                            value=int(avance_prev),
                            step=5
                        )
                 # ===== FILA 6: EQUIPO | TIPO | SEDE =====
                col7, col8, col9 = st.columns(3)
                with col7:
                    st.text_input(
                        "⚙️Equipos Rotativos / Zona de Ejecución",
                        fila["equipo"],
                        disabled=True
                    )
                with col8:
                    st.text_input(
                        "🛠 Tipo de Mantenimiento",
                        fila["tipo mantto"],
                        disabled=True
                    )
                with col9:
                    st.text_input(
                        "🏭Sede",
                        fila["sede"],
                        disabled=True
                    )

                # ===== CAMPOS SOLO PARA CORRECTIVO (PASO 1) =====
                tipo_mantto = str(fila["tipo mantto"]).strip().upper()
                causa_falla = ""
                codigo_falla = ""
                if tipo_mantto == "CORRECTIVO":
                    st.markdown("### 🔧 INFORMACIÓN DE FALLA CORRECTIVA")

                    col_cf1, col_cf2 = st.columns(2)

                    with col_cf1:
                        causa_falla = st.text_input("Causa Falla")

                    with col_cf2:
                        codigo_falla = st.text_input("Código Falla")

                 # ===== FILA FINAL =====
                continua = st.selectbox("¿Continúa?", ["Sí", "No"])
                guardar = st.form_submit_button("Guardar")
                
                if guardar:
                # ===== HORA PERÚ =====
                    tz = pytz.timezone("America/Lima")
                    hora_actual = datetime.now(tz)

                # ===== VALIDACIÓN HORAS =====
                    hi = datetime.combine(fecha_sel, hora_inicio)
                    hf = datetime.combine(fecha_sel, hora_cierre)
                    if hf <= hi:
                        st.error("La hora de cierre debe ser mayor que la de inicio")
                        st.stop()

                    duracion_final = round((hf - hi).total_seconds() / 3600, 2)
                    ws_bitacora.append_row([
                        fecha_sel.isoformat(),
                        datetime.now().strftime("%H:%M:%S"),
                        fila["pt"],
                        fila["ot"],
                        fila["equipo"],
                        fila["actividad"],
                        st.session_state.nombre,
                        detalle,
                        duracion_final,
                        avance,
                        continua,
                        st.session_state.area,
                        recurso,
                        causa_falla,
                        codigo_falla,
                        hora_inicio.strftime("%H:%M"),
                        hora_cierre.strftime("%H:%M")
                    ])

                    st.success("Registro guardado")
                    st.rerun()
# ================= ACTIVIDAD ADICIONAL =================
    with tab_adicional:

        st.subheader("➕ Registro de Actividad Adicional")

    # ===== GENERAR ID AUTOMÁTICO DESDE HOJA NUEVA =====
        df_adicional = pd.DataFrame(ws_adicional.get_all_records())

        if df_adicional.empty:
            ultimo_id = 0
        else:
            df_adicional["id_adicional"] = pd.to_numeric(df_adicional["id_adicional"], errors="coerce")
            ultimo_id = df_adicional["id_adicional"].max()

        if pd.isna(ultimo_id):
            ultimo_id = 0

        nuevo_id = str(int(ultimo_id) + 1).zfill(8)

    # ===== FORMULARIO =====
        with st.form("actividad_adicional", clear_on_submit=True):

            st.markdown(f"### 🆔 ID: {nuevo_id}")

            fecha = date.today()
            st.text_input("Fecha", fecha, disabled=True)

            actividad = st.text_area("Actividad realizada")

            detalle = st.text_area("Detalle ejecutado")

            sede = st.selectbox(
            "Sede",
            ["PGAS", "PFRAC", "TALLER MANTENIMIENTO"]
            )

            from datetime import time
            horas_turno = (
                [time(h, 0) for h in range(7, 12)] +
                [time(12, 0)] +
                [time(13, 30)] +
                [time(h, 0) for h in range(14, 20)]
            )

            col1, col2 = st.columns(2)
            with col1:
                hora_inicio = st.selectbox("Hora inicio", horas_turno)
            with col2:
                hora_cierre = st.selectbox("Hora cierre", horas_turno)

            df_users = pd.DataFrame(ws_usuarios.get_all_records())
            recursos = df_users[df_users["area"] == st.session_state.area]["Nombre"].tolist()
            recursos.insert(0, "N/A")

            recurso = st.selectbox("Recurso personal (apoyo)", recursos)

            guardar_adicional = st.form_submit_button("Guardar actividad")

            if guardar_adicional:

                hi = datetime.combine(fecha, hora_inicio)
                hf = datetime.combine(fecha, hora_cierre)

                duracion = round((hf - hi).total_seconds() / 3600, 2)

                ws_adicional.append_row([
                    fecha.isoformat(),
                    datetime.now().strftime("%H:%M:%S"),
                    nuevo_id,
                    actividad,
                    detalle,
                    st.session_state.nombre,
                    duracion,
                    st.session_state.area,
                    recurso,
                    sede,
                    hora_inicio.strftime("%H:%M"),
                    hora_cierre.strftime("%H:%M")
                ])

                st.success(f"✅ Actividad registrada con ID {nuevo_id}")
                st.rerun()
    # ================= MIS REGISTROS =================
    with tab_mis_registros:
        st.subheader("✏️ Editar Registros")
        fecha_edit = st.date_input(
        "📅 Fecha a editar",
        value=date.today()
    )
        df_bit = pd.DataFrame(ws_bitacora.get_all_records())
        df_bit["fecha"] = pd.to_datetime(df_bit["fecha"], errors="coerce").dt.date
        df_bit["duracion"] = pd.to_numeric(df_bit["duracion"], errors="coerce")
        df_bit["avance_dia"] = pd.to_numeric(df_bit["avance_dia"], errors="coerce")
        
        df_bit["mecanico"] = df_bit["mecanico"].astype(str).str.strip().str.upper()
        nombre_usuario = st.session_state.nombre.strip().upper()

        df_mios = df_bit[
            (df_bit["mecanico"] == nombre_usuario) &
            (df_bit["area"] == st.session_state.area) &
            (df_bit["fecha"] == fecha_edit)
        ]

        if df_mios.empty:
            st.info(f"📭 No tienes registros para editar el {fecha_edit.strftime('%d/%m/%Y')}")
            st.stop()
        else:
            st.dataframe(df_mios)
                # ===== SELECCIONAR REGISTRO =====
            fila_sel = st.selectbox(
                "Selecciona registro a editar",
                df_mios.index,
                format_func=lambda i: f'OT {df_mios.loc[i,"ot"]} – {df_mios.loc[i,"equipo"]}'
            )

            fila = df_mios.loc[fila_sel]

        # ===== FORMULARIO DE EDICIÓN =====
        with st.form("editar_registro"):
            detalle_edit = st.text_area(
                "Detalle ejecutado",
                fila["detalle"]
            )
            
            # ===== DURACIÓN (VALIDADA) =====

            duracion_actual = fila["duracion"]

            try:
                duracion_actual = float(duracion_actual)
            except:
                duracion_actual = 0.1
            
            if duracion_actual < 0.1:
                duracion_actual = 0.1

            duracion_edit = st.number_input(
                "Duración (h)",
                min_value=0.1,
                step=0.1,
                value=duracion_actual
            )

            avance_edit = st.slider(
                "Avance acumulado (%)",
                0, 100,
                int(fila["avance_dia"])
            )

            continua_edit = st.selectbox(
                "¿Continúa?",
                ["Sí", "No"],
                index=0 if fila["continua"] == "Sí" else 1
            )

            confirmar = st.checkbox("Confirmo que deseo modificar este registro")
            guardar_edit = st.form_submit_button("💾 Guardar cambios")

        # ===== GUARDAR EN GOOGLE SHEETS =====
        if guardar_edit:
            if not confirmar:
                st.error("Debes confirmar la modificación")
                st.stop()

            fila_sheet = fila_sel + 2  # +2 por encabezado

            ws_bitacora.update(
                f"H{fila_sheet}:K{fila_sheet}",
                [[
                    detalle_edit,
                    duracion_edit,
                    avance_edit,
                    continua_edit
                ]]
            )

            st.success("✅ Registro actualizado correctamente")
            st.rerun()
        # ================= MI AVANCE =================
    with tab_avance:
        st.subheader("📈 Avance de mis OTs")

        df_bit = pd.DataFrame(ws_bitacora.get_all_records())
        df_bit["fecha"] = pd.to_datetime(df_bit["fecha"], errors="coerce")
        df_bit["avance_dia"] = pd.to_numeric(df_bit["avance_dia"], errors="coerce")
        df_bit["duracion"] = pd.to_numeric(df_bit["duracion"], errors="coerce")

        df_mio = df_bit[
            (df_bit["mecanico"] == st.session_state.nombre) &
            (df_bit["area"] == st.session_state.area)
        ]

        if df_mio.empty:
            st.info("No hay datos suficientes para mostrar gráficos")
            st.stop()

        # ================= KPIs RÁPIDOS =================
        c1, c2, c3 = st.columns(3)
        c1.metric("OTs trabajadas", df_mio["ot"].nunique())
        c2.metric("Horas totales", round(df_mio["duracion"].sum(), 1))
        c3.metric("Avance promedio (%)", round(df_mio["avance_dia"].mean(), 1))

        st.markdown("---")

        # ================= AVANCE POR OT =================
        st.markdown("### 🔧 Avance por OT (%)")

        df_ot = (
            df_mio.groupby("ot")["avance_dia"]
            .max()
            .reset_index()
            .sort_values("avance_dia", ascending=False)
        )

        st.altair_chart(
            alt.Chart(df_ot)
            .mark_bar(size=45)
            .encode(
                x=alt.X("avance_dia:Q", title="Avance (%)"),
                y=alt.Y("ot:N", sort="-x", title="OT"),
                color=alt.Color(
                    "avance_dia:Q",
                    scale=alt.Scale(scheme="greenblue"),
                    legend=None
                ),
                tooltip=["ot", "avance_dia"]
            )
            .properties(height=max(300, 70 * len(df_ot))),
            use_container_width=True
        )

        st.markdown("---")

        # ================= TENDENCIA DE AVANCE =================
        st.markdown("### 📉 Tendencia diaria de avance")

        df_linea = (
            df_mio.groupby(df_mio["fecha"].dt.date)["avance_dia"]
            .mean()
            .reset_index()
            .rename(columns={"avance_dia": "avance_promedio"})
        )

        st.altair_chart(
            alt.Chart(df_linea)
            .mark_line(point=True)
            .encode(
                x=alt.X("fecha:T", title="Fecha"),
                y=alt.Y("avance_promedio:Q", title="Avance promedio (%)"),
                tooltip=["fecha", "avance_promedio"]
            ),
            use_container_width=True
        )
        # ================= DISTRIBUCIÓN DE AVANCE =================
        st.markdown("### 🧩 Estado de avance de mis OTs")

        def clasificar_avance(x):
            if x < 50:
                return "Bajo (<50%)"
            elif x < 95:
                return "En progreso (50–94%)"
            else:
                return "Casi terminado (≥95%)"

        df_pie = (
        df_mio.groupby("ot")["avance_dia"]
        .max()
        .reset_index()
        )

        df_pie["estado"] = df_pie["avance_dia"].apply(clasificar_avance)

        df_pie = (
        df_pie.groupby("estado")
        .size()
        .reset_index(name="cantidad")
        )

        st.altair_chart(
        alt.Chart(df_pie)
        .mark_arc(innerRadius=40)
        .encode(
        theta=alt.Theta("cantidad:Q", title="OTs"),
        color=alt.Color(
            "estado:N",
            scale=alt.Scale(
                domain=[
                    "Bajo (<50%)",
                    "En progreso (50–94%)",
                    "Casi terminado (≥95%)"
                ],
                range=["#d62728", "#ffbf00", "#2ca02c"]
            ),
            legend=alt.Legend(title="Estado de avance")
        ),
        tooltip=["estado", "cantidad"]
        ),
        use_container_width=True
        )

# ================== PALETA VISUAL DASHBOARD ==================
palette_tecnicos = [
    "#6EC1E4",  # azul claro
    "#1F77B4",  # azul
    "#FF7F0E",  # naranja
    "#2CA02C",  # verde
    "#D62728",  # rojo
    "#9467BD",  # morado
    "#17BECF",  # cyan
]
def generar_pdf(df_f):
    
    # =========================
    # ORDEN CRONOLÓGICO
    # =========================
    df_f = df_f.sort_values(by=["fecha", "ot"]).reset_index(drop=True)

    # ===== ASEGURAR COLUMNA ACTIVIDAD PARA PDF (BITÁCORA REAL) =====
    df_f.columns = df_f.columns.str.strip().str.lower()

    if "actividad_plan" in df_f.columns:
        df_f["actividad_plan"] = df_f["actividad_plan"].fillna("").astype(str)
    else:
        raise ValueError(
        "❌ La columna 'actividad_plan' NO existe en df_f. "
        "Revisa el encabezado en la hoja Bitacora."
        )

    # =========================
    # CÁLCULO DE KPIs
    # =========================
    total_registros = len(df_f)
    total_ots = df_f["ot"].nunique()
    total_personal = df_f["mecanico"].nunique()
    total_horas = round(df_f["duracion"].sum(), 1)
    avance_prom = round(df_f["avance_dia"].mean(), 1)

    # =========================
    # DOCUMENTO PDF
    # =========================
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Cell", fontSize=8, leading=10))

    story = []

    # =========================
    # TÍTULO + TEXTO INTRODUCTORIO
    # =========================
    story.append(Paragraph(
        "BITÁCORA DIGITAL DE MANTENIMIENTO – REPORTE DE CAMBIO DE GUARDIA",
        styles["Title"]
    ))
    story.append(Spacer(1, 10))

    story.append(Paragraph(
        f"""
        Estimados,<br/><br/>
        Por medio del presente se reporta el <b>cambio de guardia del área de mantenimiento</b>,
        correspondiente al periodo comprendido entre <b>{fi}</b> y <b>{ff}</b>.
        Durante este intervalo se ejecutaron las órdenes de trabajo programadas,
        registrando las actividades realizadas, horas hombre empleadas y el avance
        acumulado de cada intervención.<br/><br/>
        A continuación, se presenta el resumen ejecutivo y el detalle de las actividades
        ejecutadas, con la finalidad de asegurar la trazabilidad de los trabajos y facilitar
        la continuidad operativa del siguiente turno.
        """,
        styles["Normal"]
    ))
    # =========================
    # RESUMEN EJECUTIVO (KPIs)
    # =========================
    story.append(Spacer(1, 12))
    story.append(Paragraph("Resumen Ejecutivo del Servicio", styles["Heading2"]))
    story.append(Spacer(1, 6))

    kpi_data = [
        ["Registros", "OTs", "Personal", "Horas Totales", "Avance Promedio"],
        [
            total_registros,
            total_ots,
            total_personal,
            total_horas,
            f"{avance_prom} %"
        ]
    ]

    kpi_table = Table(kpi_data, colWidths=[4*cm]*5)
    kpi_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.6, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
    ]))

    story.append(kpi_table)
    story.append(Spacer(1, 16))
    story.append(Spacer(1, 16))

    story.append(Paragraph(f"Área: {st.session_state.area}",styles["Normal"]))
    story.append(Paragraph(f"Supervisor: {st.session_state.nombre}",styles["Normal"]))
    story.append(Paragraph(f"Periodo: {fi} al {ff}",styles["Normal"]))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",styles["Normal"]))
    story.append(Paragraph("Horas por Técnico", styles["Heading2"]))
    story.append(Spacer(1, 8))
    fig,ax=plt.subplots(figsize=(16,4),dpi=300)
    df_f.groupby("mecanico")["duracion"].sum().sort_values().plot(kind="barh",ax=ax)
    plt.tight_layout()
    img1=BytesIO()
    plt.savefig(img1,format="png",dpi=300)
    plt.close()
    img1.seek(0)
    story.append(Image(img1,width=24*cm,height=5.5*cm))
    story.append(Paragraph("OTs por Técnico", styles["Heading2"]))
    story.append(Spacer(1, 8))
    fig,ax=plt.subplots(figsize=(16,4),dpi=300)
    df_f.groupby("mecanico")["ot"].nunique().sort_values().plot(kind="barh",ax=ax)
    plt.tight_layout()
    img2=BytesIO()
    plt.savefig(img2,format="png",dpi=300)
    plt.close()
    img2.seek(0)
    story.append(Image(img2,width=24*cm,height=5.5*cm))
    story.append(Paragraph("OTs por Área", styles["Heading2"]))
    story.append(Spacer(1, 8))

    fig,ax=plt.subplots(figsize=(16,3.5),dpi=300)
    df_f.groupby("area")["ot"].nunique().sort_values().plot(kind="barh",ax=ax)
    plt.tight_layout()
    img3=BytesIO()
    plt.savefig(img3,format="png",dpi=300)
    plt.close()
    img3.seek(0)
    story.append(Image(img3,width=24*cm,height=5*cm))

    resumen=df_f.groupby("mecanico").agg(
            OTs=("ot","nunique"),
            Horas=("duracion","sum")
        ).reset_index()
    table_data=[[Paragraph(str(x),styles["Cell"]) for x in resumen.columns]]
    for _,r in resumen.iterrows():
            table_data.append([Paragraph(str(r[c]),styles["Cell"]) for c in resumen.columns])
    t=Table(table_data,colWidths=[8*cm,4*cm,4*cm])
    t.setStyle(TableStyle([
            ("GRID",(0,0),(-1,-1),0.5,colors.black),
            ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
            ("VALIGN",(0,0),(-1,-1),"TOP")
        ]))
    story.append(t)
    story.append(PageBreak())

    detalle_cols=["fecha","ot","actividad_plan","equipo","mecanico","detalle","duracion","avance_dia","continua"]
    
    detalle_headers = [
        "FECHA",
        "OT",
        "ACTIVIDAD",
        "EQUIPO/UBICACIÓN",
        "MECÁNICO",
        "DETALLE DE EJECUTADO",
        "DURACIÓN",
        "AVANCE",
        "CONTINÚA"
    ]
    data=[[Paragraph(c,styles["Cell"]) for c in detalle_headers]]

    for _,r in df_f[detalle_cols].iterrows():
            data.append([
                Paragraph(r["fecha"].strftime("%d/%m/%Y"), styles["Cell"]),
                Paragraph(str(r["ot"]),styles["Cell"]),
                Paragraph(str(r["actividad_plan"]), styles["Cell"]),   # ← NUEVO
                Paragraph(str(r["equipo"]),styles["Cell"]),
                Paragraph(str(r["mecanico"]),styles["Cell"]),
                Paragraph(str(r["detalle"]).replace("*","<br/>• "),styles["Cell"]),
                Paragraph(f'{r["duracion"]} hrs', styles["Cell"]),
                Paragraph(f'{r["avance_dia"]} %', styles["Cell"]),
                Paragraph(str(r["continua"]), styles["Cell"])   
        
            ])
    t2 = Table(
    data,
    colWidths=[
        2.0*cm,   # Fecha
        2.0*cm,   # OT
        4.5*cm,   # ACTIVIDAD
        4.0*cm,   # Equipo
        3.2*cm,   # Técnico
        7.5*cm,   # Detalle
        2.0*cm,   # Horas
        1.8*cm,   # Avance %
        1.9*cm    # Continúa
    ],
    repeatRows=1
)

    t2.setStyle(TableStyle([
            ("GRID",(0,0),(-1,-1),0.4,colors.black),
            ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
            ("VALIGN",(0,0),(-1,-1),"TOP")
        ]))
    story.append(Paragraph("Detalle de Actividades Ejecutadas",styles["Heading2"]))
    story.append(t2)
    doc.build(story)
    buffer.seek(0)
    return buffer
def generar_excel(df_f):
    # =========================
    # ORDEN CRONOLÓGICO
    # =========================
    df_f = df_f.sort_values(by=["fecha", "ot"]).reset_index(drop=True)

    # =========================
    # VALIDAR COLUMNAS (SEGURIDAD)
    # =========================
    for col in ["pt", "recurso", "hora_inicio", "hora_cierre"]:
        if col not in df_f.columns:
            df_f[col] = ""

    columnas = {
        "fecha": "Fecha",
        "ot": "OT",
        "pt": "PT",
        "equipo": "Equipo",
        "detalle": "Actividad Ejecutada",
        "mecanico": "Técnico Responsable",
        "area": "Área",
        "recurso": "Recurso Personal",
        "hora_inicio": "Hora Inicio",
        "hora_cierre": "Hora Fin",
        "duracion": "Horas Totales",
        "avance_dia": "Avance (%)",
        "continua": "Continúa"
    }

    df_export = df_f[list(columnas.keys())].copy()
    df_export["fecha"] = df_export["fecha"].dt.strftime("%d/%m/%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bitácora Cronológica"

    # Encabezados
    for col_idx, col_name in enumerate(columnas.values(), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for row_idx, row in enumerate(df_export.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    # Autoajuste de columnas
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

if st.session_state.rol in ["SUPERVISOR","PLANEAMIENTO"]:

    # ================= DASHBOARD EHS =================
    if st.session_state.area == "EHS":

        st.title("🚨 DASHBOARD DE SEGURIDAD OPERACIONAL")

        df_ots = pd.DataFrame(ws_ots.get_all_records())
        df_ots.columns = df_ots.columns.str.strip().str.lower()

        # ================= FILTRO DE FECHAS =================
        df_ots["fecha ejecucion"] = pd.to_datetime(
            df_ots["fecha ejecucion"],
            dayfirst=True,
            errors="coerce"
        )

        colf1, colf2 = st.columns(2)

        with colf1:
            fecha_inicio = st.date_input(
                "Fecha inicio",
                value=df_ots["fecha ejecucion"].min()
            )

        with colf2:
            fecha_fin = st.date_input(
                "Fecha fin",
                value=df_ots["fecha ejecucion"].max()
            )

        df_ots = df_ots[
            (df_ots["fecha ejecucion"] >= pd.to_datetime(fecha_inicio)) &
            (df_ots["fecha ejecucion"] <= pd.to_datetime(fecha_fin))
        ]

        st.markdown("---")

        col1, col2, col3, col4 = st.columns(4)

        pt_frio = (df_ots["tipo pt"] == "FRIO").sum()
        pt_caliente = (df_ots["tipo pt"] == "CALIENTE").sum()
        riesgo_alto = (df_ots["riesgo potencial"] == "ALTO").sum()
        trabajos_psm = (df_ots["psm"] == "SI").sum()

        col1.metric("PT FRÍO", pt_frio)
        col2.metric("PT CALIENTE", pt_caliente)
        col3.metric("Riesgo ALTO", riesgo_alto)
        col4.metric("Trabajos PSM", trabajos_psm)

        st.markdown("---")

        st.subheader("Distribución de Riesgo Potencial")
        riesgo = df_ots[df_ots["riesgo potencial"] != "NA"]
        riesgo = riesgo.groupby("riesgo potencial").size().reset_index(name="cantidad")
        chart_riesgo = alt.Chart(riesgo).mark_bar().encode(
            x=alt.X("riesgo potencial:N", title="Nivel de Riesgo"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("riesgo potencial:N", legend=None)
        )
        texto = alt.Chart(riesgo).mark_text(
            align="center",
            baseline="bottom",
            dy=-5,
            fontSize=14
        ).encode(
            x="riesgo potencial:N",
            y="cantidad:Q",
            text="cantidad:Q"
        )

        st.altair_chart(chart_riesgo, use_container_width=True)

        st.markdown("### Indicadores Operacionales de Seguridad")

        c1, c2, c3, c4 = st.columns(4)

        riesgo_medio = (df_ots["riesgo potencial"] == "MEDIO").sum()
        riesgo_bajo = (df_ots["riesgo potencial"] == "BAJO").sum()
        pt_total = df_ots["tipo pt"].notna().sum()
        passt = (df_ots["tipo mantto"] == "PASST").sum()

        c1.metric("Riesgo MEDIO", riesgo_medio)
        c2.metric("Riesgo BAJO", riesgo_bajo)
        c3.metric("Total PT", pt_total)
        c4.metric("Actividades PASST", passt)

        st.markdown("### Distribución de Permisos de Trabajo")
        pt_chart = df_ots[df_ots["tipo pt"] != "NA"]
        pt_chart = pt_chart.groupby("tipo pt").size().reset_index(name="cantidad")
        st.bar_chart(pt_chart.set_index("tipo pt"))

        st.markdown("### Trabajos PSM")
        psm_chart = df_ots.groupby("psm").size().reset_index(name="cantidad")
        chart_psm = alt.Chart(psm_chart).mark_bar().encode(
            x=alt.X("psm:N", title="PSM"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("psm:N", legend=None)
        )
        st.altair_chart(chart_psm, use_container_width=True)

        st.markdown("### Trabajos por Empresa")
        empresa_chart = df_ots.groupby("empresa").size().reset_index(name="cantidad")
        chart_empresa = alt.Chart(empresa_chart).mark_bar().encode(
            x=alt.X("empresa:N", title="Empresa"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("empresa:N", legend=None)
        )
        st.altair_chart(chart_empresa, use_container_width=True)

        st.markdown("### Trabajos por Sede")
        sede_chart = df_ots.groupby("sede").size().reset_index(name="cantidad")
        chart_sede = alt.Chart(sede_chart).mark_bar().encode(
            x=alt.X("sede:N", title="Sede"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("sede:N", legend=None)
        )
        st.altair_chart(chart_sede, use_container_width=True)

        st.markdown("### Tipo de Mantenimiento")
        mantto_chart = df_ots[
            (df_ots["tipo mantto"] != "GESTION") &
            (df_ots["tipo mantto"] != "PASST")
        ]
        mantto_chart = mantto_chart.groupby("tipo mantto").size().reset_index(name="cantidad")
        chart_mantto = alt.Chart(mantto_chart).mark_bar().encode(
            x=alt.X("tipo mantto:N", title="Tipo de mantenimiento"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("tipo mantto:N", legend=None)
        )
        st.altair_chart(chart_mantto, use_container_width=True)

        st.markdown("### Trabajos por Área")
        area_chart = df_ots.groupby("area").size().reset_index(name="cantidad")
        chart_area = alt.Chart(area_chart).mark_bar().encode(
            x=alt.X("area:N", title="Área"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("area:N", legend=None)
        )

        st.altair_chart(chart_area, use_container_width=True)

        st.markdown("### Índice de Riesgo Operacional")
        total_trabajos = len(df_ots)
        if total_trabajos > 0:
            indice_riesgo = round((riesgo_alto / total_trabajos) * 100, 2)
        else:
            indice_riesgo = 0
        pt_caliente_pct = round((pt_caliente / pt_total) * 100, 2) if pt_total > 0 else 0
        psm_pct = round((trabajos_psm / total_trabajos) * 100, 2) if total_trabajos > 0 else 0
        riesgo_reducido = (df_ots["riesgo controles"] != df_ots["riesgo potencial"]).sum()
        indice_control = round((riesgo_reducido / total_trabajos) * 100, 2) if total_trabajos > 0 else 0

        # ===== Mostrar KPIs en columnas =====
        c1, c2, c3, c4 = st.columns(4)

        c1.metric("Índice de Riesgo (%)", f"{indice_riesgo}%")
        c2.metric("Índice PT Caliente (%)", f"{pt_caliente_pct}%")
        c3.metric("Índice PSM (%)", f"{psm_pct}%")
        c4.metric("Riesgos Controlados (%)", f"{indice_control}%")

        st.markdown("### Riesgos Altos por Área")
        riesgo_area = (
            df_ots[df_ots["riesgo potencial"] == "ALTO"]
            .groupby("area")
            .size()
        )
        riesgo_area = (
            df_ots[df_ots["riesgo potencial"] == "ALTO"]
            .groupby("area")
            .size()
            .reset_index(name="cantidad")
        )
        chart_riesgo_area = alt.Chart(riesgo_area).mark_bar().encode(
            x=alt.X("area:N", title="Área"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("area:N", legend=None)
        )
        st.altair_chart(chart_riesgo_area, use_container_width=True)

        st.markdown("### Riesgo Alto por Sede")
        riesgo_sede = (
            df_ots[df_ots["riesgo potencial"] == "ALTO"]
            .groupby("sede")
            .size()
        )
        riesgo_sede = (
            df_ots[df_ots["riesgo potencial"] == "ALTO"]
            .groupby("sede")
            .size()
            .reset_index(name="cantidad")
        )

        chart_riesgo_sede = alt.Chart(riesgo_sede).mark_bar().encode(
            x=alt.X("sede:N", title="Sede"),
            y=alt.Y("cantidad:Q", title="Cantidad"),
            color=alt.Color("sede:N", legend=None)
        )

        st.altair_chart(chart_riesgo_sede, use_container_width=True)
        
        st.markdown("### Tendencia de Trabajos por Día")
        trabajos_dia = (
            df_ots.groupby("fecha ejecucion")
            .size()
        )
        st.line_chart(trabajos_dia)

        st.markdown("### Tendencia de Riesgo Alto")

        riesgo_dia = (
            df_ots[df_ots["riesgo potencial"] == "ALTO"]
            .groupby("fecha ejecucion")
            .size()
        )
        st.line_chart(riesgo_dia)

        passt_pct = round((passt / total_trabajos) * 100, 2) if total_trabajos > 0 else 0
        st.metric("Actividades PASST (%)", f"{passt_pct}%")
        
        exposicion = riesgo_alto + pt_caliente + trabajos_psm
        indice_exposicion = round((exposicion / total_trabajos) * 100, 2) if total_trabajos > 0 else 0
        st.metric("Índice de Exposición Operacional", f"{indice_exposicion}%")

        st.stop()


    st.title(f"📊 {st.session_state.cargo} – KPIs")
    df=pd.DataFrame(ws_bitacora.get_all_records())
    df["fecha"]=pd.to_datetime(df["fecha"],errors="coerce")
    df["duracion"]=pd.to_numeric(df["duracion"],errors="coerce")
    df["avance_dia"]=pd.to_numeric(df["avance_dia"],errors="coerce")
    df=df.dropna(subset=["fecha"])
    if st.session_state.rol=="SUPERVISOR":
        df=df[df["area"]==st.session_state.area]
    fi=st.date_input("Inicio",value=df["fecha"].min().date())
    ff=st.date_input("Fin",value=df["fecha"].max().date())
    df_f = df[
        (df["fecha"] >= pd.to_datetime(fi)) &
        (df["fecha"] <= pd.to_datetime(ff))
    ].copy()
    # asegurar orden cronológico también fuera del PDF
    df_f = df_f.sort_values(by=["fecha", "ot"]).reset_index(drop=True)

    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric("Registros",len(df_f))
    c2.metric("OTs",df_f["ot"].nunique())
    c3.metric("Personal",df_f["mecanico"].nunique())
    c4.metric("Horas",round(df_f["duracion"].sum(),1))
    c5.metric("Avance %",round(df_f["avance_dia"].mean(),1))

    st.subheader("OTs por Área")
    df_area=df_f.groupby("area")["ot"].nunique().reset_index()
    h_area=max(200,80*len(df_area))
    st.altair_chart(
        alt.Chart(df_area)
        .mark_bar(size=50)
        .encode(
            x=alt.X("ot:Q",title="OTs"),
            y=alt.Y("area:N",title="Área",sort="-x"),
            color=alt.Color("area:N",scale=alt.Scale(range=palette_tecnicos),legend=alt.Legend(title="Área")),
            tooltip=["area","ot"]
        )
        .properties(height=h_area),
        use_container_width=True
    )

    st.markdown("---")

    st.subheader("Horas por Técnico")
    df_tec=df_f.groupby("mecanico")["duracion"].sum().reset_index()
    h_tec=max(300,80*len(df_tec))
    st.altair_chart(
        alt.Chart(df_tec)
        .mark_bar(size=45)
        .encode(
            x=alt.X("duracion:Q",title="Horas"),
            y=alt.Y("mecanico:N",title="Técnico",sort="-x"),
            color=alt.Color("mecanico:N",scale=alt.Scale(range=palette_tecnicos),legend=alt.Legend(title="Técnico")),
            tooltip=["mecanico",alt.Tooltip("duracion",format=".1f")]
        )
        .properties(height=h_tec),
        use_container_width=True
    )

    st.markdown("---")

    st.subheader("OTs por Técnico")
    df_tec_ot=df_f.groupby("mecanico")["ot"].nunique().reset_index()
    h_tec_ot=max(300,80*len(df_tec_ot))
    st.altair_chart(
        alt.Chart(df_tec_ot)
        .mark_bar(size=45)
        .encode(
            x=alt.X("ot:Q",title="Cantidad de OTs"),
            y=alt.Y("mecanico:N",title="Técnico",sort="-x"),
            color=alt.Color("mecanico:N",scale=alt.Scale(range=palette_tecnicos),legend=alt.Legend(title="Técnico")),
            tooltip=["mecanico","ot"]
        )
        .properties(height=h_tec_ot),
        use_container_width=True
    )

    st.dataframe(df_f)

    # ============================================================
# 🔥 POWER UP: PERFIL DE TÉCNICO (ANÁLISIS INDIVIDUAL)
# ============================================================

    st.markdown("---")
    st.markdown("## 👨‍🔧 Análisis de Desempeño por Técnico")

# ================= SELECTOR =================
    df_users = pd.DataFrame(ws_usuarios.get_all_records())
    tecnicos = df_f["mecanico"].dropna().unique().tolist()

    if len(tecnicos) == 0:
        st.warning("No hay técnicos disponibles en el rango seleccionado")
    else:
        tecnico_sel = st.selectbox("Seleccionar técnico", tecnicos)

    # ================= FILTRO =================
        df_tec = df_f[df_f["mecanico"] == tecnico_sel]

    # ================= FOTO + KPIs =================
        col1, col2 = st.columns([1,3])

        fila_user = df_users[df_users["Nombre"] == tecnico_sel]
    
    if not fila_user.empty:
        usuario_codigo = fila_user.iloc[0]["Usuario"]
        ruta_foto = f"fotos/{usuario_codigo}.jpg"
    else:
        ruta_foto = "fotos/default.jpg"

    ruta_default = "fotos/default.jpg"
    
    with col1:
        if os.path.exists(ruta_foto):
            st.image(ruta_foto, width=150)
        else:
            st.image(ruta_default, width=150)

    with col2:
        st.subheader(tecnico_sel)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("OTs ejecutadas", df_tec["ot"].nunique())
        c2.metric("Horas trabajadas", round(df_tec["duracion"].sum(),1))
        c3.metric("Avance promedio (%)", round(df_tec["avance_dia"].mean(),1))
        c4.metric("Registros", len(df_tec))

    st.markdown("---")

    # ================= EQUIPOS =================
    st.markdown("### ⚙️ Equipos intervenidos")

    df_eq = df_tec["equipo"].value_counts().reset_index()
    df_eq.columns = ["equipo","intervenciones"]

    st.bar_chart(df_eq.set_index("equipo"))

    st.markdown("---")

    # ================= HORAS POR DÍA =================
    st.markdown("### ⏱ Horas trabajadas por día")

    df_horas = (
        df_tec.groupby(df_tec["fecha"].dt.date)["duracion"]
        .sum()
        .reset_index()
    )

    st.line_chart(df_horas.set_index("fecha"))

    st.markdown("---")

    # ================= OTs =================
    st.markdown("### 🧾 OTs ejecutadas")

    df_ot = (
        df_tec.groupby("ot")["duracion"]
        .sum()
        .reset_index()
        .sort_values("duracion", ascending=False)
    )

    st.bar_chart(df_ot.set_index("ot"))

    st.markdown("---")

    # ================= HISTORIAL =================
    st.markdown("### 📋 Historial detallado")

    st.dataframe(
        df_tec.sort_values(by="fecha", ascending=False),
        use_container_width=True
    )

    pdf = generar_pdf(df_f)
    st.download_button(
        "📄 Exportar Cambio de Guardia (PDF)",
        pdf,
        file_name="Cambio_Guardia.pdf",
        mime="application/pdf"
    )

    excel = generar_excel(df_f)
    st.download_button(
    "📊 Exportar Bitácora Cronológica (Excel)",
        excel,
        file_name="Bitacora_Cronologica.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
